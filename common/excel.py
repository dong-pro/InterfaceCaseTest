# -*- coding: utf-8 -*-
from openpyxl import load_workbook
import time, datetime


class Excel:

    @classmethod
    def deleteRows(self, filePath, start, count):
        wb = load_workbook(filePath)
        ws = wb.active
        ws.delete_rows(start, count)  # 删除index为start后面的count行
        wb.save(filePath)

    @classmethod
    def addRows(self, filePath, rows):
        wb = load_workbook(filePath)
        sheet = wb.get_sheet_by_name("Export")
        for row in rows:
            sheet.append(row)
        wb.save(filePath)

    @classmethod
    def update(self, filePath, coordinate, value):
        filePath = filePath
        wb = load_workbook(filePath)
        ws = wb.active
        ws[coordinate] = value
        wb.save(filePath)

    # 获取发票数据行数
    @classmethod
    def getInvoiceLines(self, filePath):
        print(filePath)
        wb = load_workbook(filePath, data_only=True)
        ws = wb.get_sheet_by_name("发票明细")
        num = 3
        # 单据列表
        list = []
        while 1:
            # 供应商名称
            sellerName = ws.cell(row=num, column=1).value
            buyerName = ws.cell(row=num, column=2).value
            amount = ws.cell(row=num, column=7).value
            startdate = ws.cell(row=num, column=10).value

            if sellerName == None:
                return list
            else:
                num = num + 1
            info = {}
            info['sellerName'] = sellerName
            info['buyerName'] = buyerName
            info['amount'] = amount
            ss = startdate.strftime('%Y-%m-%d')
            info['startdate'] = ss
            list.append(info)

    # 获取应收数据行数
    @classmethod
    def getReceiveLines(self, filePath):
        print(filePath)
        wb = load_workbook(filePath, data_only=True)
        ws = wb.get_sheet_by_name("应收账款")
        num = 3
        # 单据列表
        list = []
        while 1:
            # 供应商名称
            sellerName = ws.cell(row=num, column=1).value
            buyerName = ws.cell(row=num, column=2).value
            amount = ws.cell(row=num, column=4).value
            startdate = ws.cell(row=num, column=3).value
            period = ws.cell(row=num, column=5).value
            if sellerName == None:
                return list
            else:
                num = num + 1
            info = {}
            info['sellerName'] = sellerName
            info['buyerName'] = buyerName
            info['amount'] = amount
            ss = (startdate + datetime.timedelta(days=-period)).strftime('%Y-%m-%d')
            info['startdate'] = ss
            info['period'] = period
            list.append(info)

    # 获取回款数据行数
    @classmethod
    def getBackLines(self, filePath):
        print(filePath)
        wb = load_workbook(filePath, data_only=True)
        ws = wb.get_sheet_by_name("付款数据")
        num = 3
        # 单据列表
        list = []
        while 1:
            # 供应商名称
            sellerName = ws.cell(row=num, column=1).value
            buyerName = ws.cell(row=num, column=2).value
            amount = ws.cell(row=num, column=7).value
            startdate = ws.cell(row=num, column=8).value
            if sellerName == None:
                return list
            else:
                num = num + 1
            info = {}
            info['sellerName'] = sellerName
            info['buyerName'] = buyerName
            info['amount'] = amount
            # info['startdate'] = startdate
            ss = startdate.strftime('%Y-%m-%d')
            info['startdate'] = ss
            list.append(info)
